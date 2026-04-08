import os
from openpyxl import load_workbook
from datetime import datetime, time, timedelta
import re
import io
from dotenv import load_dotenv
from smbclient import open_file, register_session

load_dotenv('key.env')

SMB_USER = os.getenv("SMB_USER")
SMB_USER_PASSWORD = os.getenv("SMB_USER_PASSWORD")
FOLDER = os.getenv("FOLDER")

register_session(
    FOLDER,
    username = SMB_USER,
    password = SMB_USER_PASSWORD,
    auth_protocol ='ntlm'
    )
year = datetime.now().year
FILE_PATH = f'график дежурств { year }.xlsx'
SHARE_PATH = rf"\\{FOLDER}\IT\{FILE_PATH}"



MONTH = {
    "Январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12
}

USER_DUTY = {
    "Гриченко": "Гриченко Павел",
    "Залецкий": "Залецкий Евгений",
    "Шутов": "Шутов Алексей",
}

def day_to_duty():
    now = datetime.now()
    current_date = now.day
    current_month = now.month
    message = "Сегодня"

    # Если время больше 21:30, переключаемся на "завтра"
    if time(21, 30) <= now.time() <= time(23, 59):
        next_day = now + timedelta(days=1)
        message = "Завтра"
        current_date = next_day.day
        current_month = next_day.month

    # Получаем название месяца
    actual_month = next((month for month, num in MONTH.items() if num == current_month), None)
    if actual_month is None:
        return "Ошибка: месяц не найден."

    # Открываем книгу и выбираем лист
    try:
        with open_file(SHARE_PATH, mode = 'rb') as file:
            file_data = io.BytesIO(file.read())
        wb = load_workbook(file_data, data_only=True)
        if actual_month not in wb.sheetnames:
            return f"Ошибка: Лист '{ actual_month }' не найден в файле"
        sheet = wb[actual_month]
    except Exception as e:
        return f"Ошибка при загрузке файла: {e}"

    # Поиск даты в первой строке
    column_letter = None
    for row in sheet['A1:AF1']:
        for cell in row:
            if cell.value == current_date:
                column_letter = re.sub(r'\d', '', cell.coordinate)
                break
        if column_letter:
            break

    if not column_letter:
        return f"Ошибка: дата {current_date} не найдена в файле."

    # Поиск дежурного
    duty_row = None
    for cell in sheet[column_letter]:
        if str(cell.value).strip().lower() in ['x', 'х']:
            duty_row = re.sub(r'\D', '', cell.coordinate)
            break

    if not duty_row:
        return f"Ошибка: не найден дежурный на {message.lower()}."

    # Получение фамилии дежурного
    duty_person = None
    for row in sheet.iter_rows(min_row=int(duty_row), max_row=int(duty_row), values_only=True):
        surname = row[0]  # Первая колонка содержит фамилию
        if surname in USER_DUTY:
            duty_person = USER_DUTY[surname]
            break

    if not duty_person:
        return f"Ошибка: дежурный не найден в списке сотрудников."
    print(f"{message} дежурит: <b>{duty_person}</b>")
    return f"{message} дежурит: <b>{duty_person}</b>"
